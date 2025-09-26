import os, re, sqlite3, datetime, secrets, io, json
from functools import wraps

from flask import (
    Flask, request, redirect, url_for, session, render_template_string,
    flash, send_from_directory, send_file, Response
)

from openpyxl import load_workbook, Workbook
from urllib.parse import urlencode

# Security & CSRF
from werkzeug.security import generate_password_hash, check_password_hash
from flask_wtf import CSRFProtect
from flask_wtf.csrf import generate_csrf

from markupsafe import escape
def h(x):
    """HTML-escape helper for building safe HTML strings."""
    return '' if x is None else str(escape(x))


# ------------------------- Email (SendGrid, optional) -------------------------
def send_email(to, subject, html):
    """Best-effort email. Safe if SENDGRID not configured."""
    try:
        from sendgrid import SendGridAPIClient
        from sendgrid.helpers.mail import Mail
        key = os.environ.get("SENDGRID_API_KEY")
        frm = os.environ.get("MAIL_FROM")
        if not key or not frm or not to:
            return
        sg = SendGridAPIClient(key)
        msg = Mail(from_email=frm, to_emails=to, subject=subject, html_content=html)
        sg.send(msg)
    except Exception:
        # Silent: email is optional
        pass

# ------------------------------ App constants --------------------------------
BUILD_TAG = "HMS-2025-09-24-fixes-r5"

APP_TITLE = "Hiring Management System (HMS)"
BASE_DIR = os.path.dirname(__file__)
DB_PATH = "/home/dcdchiringsystem/DCDC-Hiring/hms.db"
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Safer default so app boots even if env var missing; set real secret on Web tab
_env_secret = os.environ.get("HMS_SECRET") or os.environ.get("SECRET_KEY")

if _env_secret:
    SECRET_KEY = _env_secret
else:
    # Only allow a fallback when running the local dev server
    if __name__ == "__main__":  # python app.py
        SECRET_KEY = "dev-only-change-me"
        print("[HMS] DEV: using fallback SECRET_KEY; DO NOT use in production.")
    else:
        raise RuntimeError("HMS_SECRET env var is required in production.")

LOGO_FILENAME = "logo.png"
POSTS = [
    "Trainee", "Junior Technician", "Senior Technician",
    "Staff Nurse", "Doctor", "DMO", "Others"
]

ROLE_ADMIN="admin"; ROLE_VP="vp"; ROLE_HR="hr"; ROLE_MANAGER="manager"; ROLE_INTERVIEWER="interviewer"
ALLOWED_CV_EXTS = {".pdf",".doc",".docx"}

# ------------------------------- Flask + CSRF --------------------------------
app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config.update(
    MAX_CONTENT_LENGTH=16 * 1024 * 1024,  # 16MB
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
)
# In production behind HTTPS, uncomment:
if os.environ.get("FLASK_ENV") == "production" or not app.debug:
    app.config["SESSION_COOKIE_SECURE"] = True

csrf = CSRFProtect(app)
app.config["WTF_CSRF_SSL_STRICT"] = False
@app.after_request
def add_security_headers(resp):
    # Clickjacking protection
    resp.headers.setdefault("X-Frame-Options", "DENY")
    # Prevent MIME type sniffing
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    # Don’t send referrer to other sites
    resp.headers.setdefault("Referrer-Policy", "no-referrer")
    # Avoid caching for authenticated users
    if 'user_id' in session:
        resp.headers["Cache-Control"] = "no-store"
    return resp

@app.context_processor
def inject_csrf():
    # For templates that want to call {{ csrf_token() }} manually
    return dict(csrf_token=generate_csrf)

# --------------------------------- Database ----------------------------------
def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    # Enforce FK constraints on every connection
    try:
        conn.execute("PRAGMA foreign_keys=ON")
    except Exception:
        pass
    return conn

def ensure_column(table: str, column: str, decl: str):
    """Add a column if it does not exist (SQLite)."""
    db = get_db(); cur = db.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = {row[1] for row in cur.fetchall()}
    if column not in cols:
        cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {decl}")
        db.commit()
    db.close()

def ensure_index(sql: str):
    db = get_db(); cur = db.cursor()
    cur.execute(sql)
    db.commit(); db.close()

def init_db():
    conn = get_db(); c = conn.cursor()
    if os.environ.get("SQLITE_JOURNAL", "").upper() == "WAL":
        try:
            conn.execute("PRAGMA journal_mode=WAL")
        except Exception:
            pass

    
    c.execute("""
    CREATE TABLE IF NOT EXISTS users(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      name TEXT NOT NULL,
      email TEXT UNIQUE NOT NULL,
      role TEXT NOT NULL,
      manager_id INTEGER,
      passcode TEXT NOT NULL,
      created_at TEXT NOT NULL
    );""")

    c.execute("""
    CREATE TABLE IF NOT EXISTS password_resets(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      user_email TEXT NOT NULL,
      state TEXT NOT NULL,
      created_at TEXT NOT NULL,
      resolved_at TEXT,
      resolver_id INTEGER,
      new_passcode TEXT
    );""")

    c.execute("""
    CREATE TABLE IF NOT EXISTS candidates(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      candidate_code TEXT,
      salutation TEXT,
      full_name TEXT NOT NULL,
      email TEXT,
      qualification TEXT,
      experience_years REAL,
      current_designation TEXT,
      phone TEXT,
      cv_path TEXT,
      current_salary TEXT,
      expected_salary TEXT,
      current_location TEXT,
      preferred_location TEXT,
      post_applied TEXT NOT NULL,
      interview_date TEXT,
      current_previous_company TEXT,
      assigned_region TEXT,
      status TEXT NOT NULL,
      decision_by INTEGER,
      remarks TEXT,
      created_by INTEGER NOT NULL,
      created_at TEXT NOT NULL,
      interviewer_id INTEGER,
      manager_owner INTEGER,
      final_decision TEXT,
      final_remark TEXT,
      finalized_by INTEGER,
      finalized_at TEXT,
      hr_join_status TEXT,
      hr_joined_at TEXT
    );""")

    c.execute("""
    CREATE TABLE IF NOT EXISTS interviews(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      candidate_id INTEGER NOT NULL,
      interviewer_id INTEGER NOT NULL,
      feedback TEXT,
      rating INTEGER,
      decision TEXT,
      is_reinterview INTEGER DEFAULT 0,
      is_edit INTEGER DEFAULT 0,
      edited_from INTEGER,
      created_at TEXT NOT NULL
    );""")

    c.execute("""
    CREATE TABLE IF NOT EXISTS notifications(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      user_id INTEGER NOT NULL,
      title TEXT NOT NULL,
      body TEXT,
      is_read INTEGER DEFAULT 0,
      created_at TEXT NOT NULL
    );""")

    # Seed first-time users
    c.execute("SELECT COUNT(*) AS ct FROM users")
    if (c.fetchone()["ct"] or 0) == 0:
        now = datetime.datetime.utcnow().isoformat()
        seed = [
            ("Mr. Parveen Chaudhary","clinicalanalyst@dcdc.co.in",ROLE_ADMIN,None),
            ("Mr. Deepak Agarwal","drdeepak@dcdc.co.in",ROLE_VP,None),
            ("Ms. Barkha","jobs@dcdc.co.in",ROLE_HR,None),
            ("Deepika","hiring@dcdc.co.in",ROLE_HR,None),
            ("Karishma","hr_hiring@dcdc.co.in",ROLE_HR,None),
            ("Kajal","hiring_1@dcdc.co.in",ROLE_HR,None),
            ("Sneha","hiring_2@dcdc.co.in",ROLE_HR,None),
            ("Ravi","hiring_3@dcdc.co.in",ROLE_HR,None),
            ("Shivani","recruitments@dcdc.co.in",ROLE_HR,None),
            ("Udita","careers@dcdc.co.in",ROLE_HR,None),
            ("Dr. Yasir Anis","clinical_manager@dcdc.co.in",ROLE_MANAGER,None),
            ("Ms. Prachi","infectioncontroller@dcdc.co.in",ROLE_INTERVIEWER,None),
            ("Mr. Shaikh Saadi","dialysis.coord@dcdc.co.in",ROLE_MANAGER,None),
            ("Ms. Pankaja","rmclinical_4@dcdc.co.in",ROLE_INTERVIEWER,None),
            ("Mr. Yekula Bhanu Prakash","rmclinical_6@dcdc.co.in",ROLE_INTERVIEWER,None),
            ("Mr. Rohit","clinical_therapist@dcdc.co.in",ROLE_INTERVIEWER,None),
        ]
        seeded_credentials = []
        for n,e,r,m in seed:
            temp_pass = secrets.token_urlsafe(8)
            c.execute("INSERT INTO users(name,email,role,manager_id,passcode,created_at) VALUES(?,?,?,?,?,?)",
                      (n,e,r,m,generate_password_hash(temp_pass),now))
            seeded_credentials.append((e, temp_pass))

        # Link interviewers to managers
        def uid(em):
            c.execute("SELECT id FROM users WHERE email=?", (em,))
            rr = c.fetchone(); return rr["id"] if rr else None
        yasir = uid("clinical_manager@dcdc.co.in")
        saadi = uid("dialysis.coord@dcdc.co.in")
        c.execute("UPDATE users SET manager_id=? WHERE email='infectioncontroller@dcdc.co.in'", (yasir,))
        for em in ("rmclinical_4@dcdc.co.in","rmclinical_6@dcdc.co.in","clinical_therapist@dcdc.co.in"):
            c.execute("UPDATE users SET manager_id=? WHERE email=?", (saadi, em))
        if seeded_credentials:
            print("[HMS] Seeded default users with temporary passcodes:")
            for email, passcode in seeded_credentials:
                print("   - {} : {}".format(email, passcode))
    conn.commit(); conn.close()

    # Ensure new columns exist for older databases
    try:
        ensure_column('password_resets', 'token', 'TEXT')
        ensure_column('password_resets', 'expires_at', 'TEXT')
    except Exception:
        pass

    # Helpful indexes (no-ops if already exist)
    try:
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_created_at ON candidates(created_at)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_status ON candidates(status)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_final_decision ON candidates(final_decision)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_hr_join_status ON candidates(hr_join_status)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_post_applied ON candidates(post_applied)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_assigned_region ON candidates(assigned_region)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_created_by ON candidates(created_by)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_manager_owner ON candidates(manager_owner)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_interviewer_id ON candidates(interviewer_id)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_notifications_user ON notifications(user_id, is_read, created_at)")
    except Exception:
        pass

def migrate_db():
    """Add missing columns if DB created from an older build."""
    conn = get_db(); c = conn.cursor()

    # password_resets: token & expires_at for email flow
    c.execute("PRAGMA table_info(password_resets)")
    cols = {row[1] for row in c.fetchall()}
    if "token" not in cols:
        c.execute("ALTER TABLE password_resets ADD COLUMN token TEXT")
    if "expires_at" not in cols:
        c.execute("ALTER TABLE password_resets ADD COLUMN expires_at TEXT")

    # interviews: edit audit trail
    c.execute("PRAGMA table_info(interviews)")
    icols = {row[1] for row in c.fetchall()}
    if "is_edit" not in icols:
        c.execute("ALTER TABLE interviews ADD COLUMN is_edit INTEGER DEFAULT 0")
    if "edited_from" not in icols:
        c.execute("ALTER TABLE interviews ADD COLUMN edited_from INTEGER")

    conn.commit(); conn.close()

    # Also ensure indexes each boot
    try:
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_created_at ON candidates(created_at)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_status ON candidates(status)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_final_decision ON candidates(final_decision)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_hr_join_status ON candidates(hr_join_status)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_post_applied ON candidates(post_applied)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_assigned_region ON candidates(assigned_region)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_created_by ON candidates(created_by)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_manager_owner ON candidates(manager_owner)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_candidates_interviewer_id ON candidates(interviewer_id)")
        ensure_index("CREATE INDEX IF NOT EXISTS idx_notifications_user ON notifications(user_id, is_read, created_at)")
    except Exception:
        pass

# Initialize/migrate DB at import time (on uWSGI load)
try:
    init_db(); migrate_db()
    print("[HMS] DB initialized/migrated", BUILD_TAG)
except Exception as e:
    print("[HMS] DB init/migrate warning:", e)

# --------------------------- Helper / auth utilities --------------------------
def current_user():
    uid = session.get("user_id")
    if not uid: return None
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT * FROM users WHERE id=?", (uid,))
    row = cur.fetchone(); db.close()
    return row

def is_hr_head(u): return u and u["role"]==ROLE_HR and u["email"].lower()=="jobs@dcdc.co.in"

def login_required(f):
    @wraps(f)
    def w(*a,**kw):
        if not current_user(): return redirect(url_for("login"))
        return f(*a,**kw)
    return w

def role_required(*roles):
    def deco(f):
        @wraps(f)
        def w(*a,**kw):
            u = current_user()
            if not u or u["role"] not in roles:
                flash("You do not have permission.","error")
                return redirect(url_for("dashboard"))
            return f(*a,**kw)
        return w
    return deco

def user_id_by_email(email:str):
    db=get_db(); cur=db.cursor()
    cur.execute("SELECT id FROM users WHERE email=?", (email,))
    r=cur.fetchone(); db.close()
    return r["id"] if r else None

def user_email_by_id(uid:int):
    db=get_db(); cur=db.cursor()
    cur.execute("SELECT email FROM users WHERE id=?", (uid,))
    r=cur.fetchone(); db.close()
    return r["email"] if r else None

def interviewers_for_manager(mid:int):
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT id,name FROM users WHERE role='interviewer' AND manager_id=? ORDER BY name", (mid,))
    rows = cur.fetchall(); db.close()
    return rows

def all_interviewers():
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT id,name FROM users WHERE role='interviewer' ORDER BY name")
    rows = cur.fetchall(); db.close()
    return rows

def notify(user_id:int, title:str, body:str=""):
    db=get_db(); cur=db.cursor()
    cur.execute("INSERT INTO notifications(user_id,title,body,is_read,created_at) VALUES(?,?,?,?,?)",
                (user_id, title, body, 0, datetime.datetime.utcnow().isoformat()))
    db.commit(); db.close()
    em = user_email_by_id(user_id)
    if em:
        send_email(em, title, "<pre style='white-space:pre-wrap'>{}</pre>".format(body))

def next_candidate_code():
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT MAX(id) FROM candidates")
    row = cur.fetchone(); db.close()
    return "DCDC_C{}".format((row[0] or 0) + 1)

# unread notifications badge
@app.context_processor
def inject_unread():
    u = current_user()
    n = 0
    if u:
        db=get_db(); cur=db.cursor()
        cur.execute("SELECT COUNT(*) FROM notifications WHERE user_id=? AND is_read=0", (u["id"],))
        n = cur.fetchone()[0] or 0
        db.close()
    return dict(unread_notifications=n)

# ------------------------------ Base HTML layout ------------------------------
BASE_HTML = """
<!doctype html>
<html>
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{{ title }}</title>
<style>
:root { --vein-blue:#0b5394; --artery-red:#b91c1c; --muted:#e5e7eb; --text:#0f172a; }
body{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;margin:0;background:#fafafa;color:var(--text)}
header{background:var(--vein-blue);color:#fff;padding:10px 12px;display:flex;gap:12px;align-items:center;flex-wrap:wrap}
header a{color:#e2e8f0;text-decoration:none;margin-right:12px}
.brand{font-weight:800;display:flex;align-items:center;gap:10px}
.brand img{height:28px;width:auto;display:block;border-radius:6px;background:#fff}
.wrap{max-width:1100px;margin:14px auto;padding:0 12px}
.card{background:#fff;border:1px solid var(--muted);border-radius:14px;padding:14px;margin:12px 0;box-shadow:0 1px 2px rgba(0,0,0,.06)}
.row{display:flex;flex-wrap:wrap;gap:12px}
.col{flex:1;min-width:280px}
.btn{display:inline-block;padding:8px 12px;border-radius:10px;border:1px solid var(--vein-blue);background:var(--vein-blue);color:#fff;text-decoration:none;cursor:pointer}
.btn.light{background:#fff;color:var(--vein-blue)}
.btn.warn{background:#fff;color:#b45309;border-color:#b45309}
.btn.danger{background:var(--artery-red);border-color:var(--artery-red)}
input,select,textarea{width:100%;padding:10px;border:1px solid #d1d5db;border-radius:10px;margin-top:6px}
label{font-weight:600}
table{width:100%;border-collapse:collapse}
th,td{padding:8px;border-bottom:1px solid #eee;text-align:left}
.tag{display:inline-block;padding:2px 8px;border-radius:999px;font-size:12px;border:1px solid var(--muted);background:#f8fafc}
.flash{padding:8px 12px;border-radius:10px;margin:8px 0}
.ok{background:#ecfdf5;border:1px solid #10b981}
.error{background:#fef2f2;border:1px solid var(--artery-red)}
.nav{display:flex;align-items:center;flex-wrap:wrap}
.nav a{margin-right:10px}
.form-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:10px}
.badge{display:inline-block;padding:4px 10px;border-radius:999px;background:#eef6ff;border:1px solid #dbeafe;color:#0b5394}
.grid-2{display:grid;grid-template-columns:1fr 1fr;gap:14px}
@media (max-width:880px){ .grid-2{grid-template-columns:1fr} }
.field{display:flex;flex-direction:column;gap:6px}
.field small.hint{color:#64748b;font-size:12px;margin-top:-4px}
.req::after{content:" *"; color:#b91c1c; font-weight:700}
.sticky-actions{position:sticky;bottom:10px;display:flex;gap:10px;justify-content:flex-end;padding-top:10px}
.card.section{padding:16px 16px 10px;border-left:6px solid #e5e7eb}
.card.section h4{margin:0 0 8px 0}
.section.blue{border-left-color:#0b53941a}
.chip{display:inline-block;padding:4px 10px;border-radius:999px;background:#eef6ff;border:1px solid #dbeafe;color:#0b5394;margin-right:6px}
.nav .bell-link{position:relative;display:inline-flex;align-items:center;justify-content:center;text-decoration:none;font-size:18px;margin-left:4px}
.nav .bell-badge{position:absolute;top:-6px;right:-10px;background:#ef4444;color:#fff;border-radius:999px;padding:0 6px;font-size:12px;border:2px solid #fff;line-height:18px;min-width:18px;text-align:center}
</style>
</head>
<body>
<header>
  <div class="brand">
    <img src="{{ url_for('brand_logo') }}" alt="logo" onerror="this.style.display='none'">
    <div>{{ app_title }}</div>
  </div>
  <nav class="nav">
    {% if user %}
      <a href="{{ url_for('dashboard') }}">Dashboard</a>
      <a href="{{ url_for('candidates_all') }}">Candidates</a>
      {% if user['role'] in ['hr','admin'] %}
        <a href="{{ url_for('add_candidate') }}">Add Candidate</a>
        <a href="{{ url_for('bulk_upload') }}">Bulk Upload</a>
        <a href="{{ url_for('hr_join_queue') }}">HR Actions</a>
      {% endif %}
      {% if user['role'] in ['admin'] %}
        <a href="{{ url_for('admin_users') }}">Admin</a>
      {% endif %}
      <a href="{{ url_for('profile') }}">Profile</a>
      <a href="{{ url_for('logout') }}">Logout</a>
      <a href="{{ url_for('notifications') }}" class="bell-link" title="Notifications" aria-label="Notifications">🔔
        {% if (unread_notifications or 0)|int > 0 %}
          <span class="bell-badge">{{ unread_notifications }}</span>
        {% endif %}
      </a>
    {% else %}
      <a href="{{ url_for('login') }}">Login</a>
    {% endif %}
  </nav>
</header>
<div class="wrap">
{% with messages = get_flashed_messages(with_categories=true) %}
  {% if messages %}
    {% for cat, msg in messages %}
      <div class="flash {{ 'ok' if cat=='message' else cat }}">{{ msg }}</div>
    {% endfor %}
  {% endif %}
{% endwith %}
{{ body|safe }}
</div>
</body>
</html>
"""

def render_page(title, body_html):
    return render_template_string(BASE_HTML, title=title, app_title=APP_TITLE, user=current_user(), body=body_html)

# ------------------------------ Misc small routes -----------------------------
@app.route("/__version")
def __version():
    return "HMS build: {}".format(BUILD_TAG)

@app.route("/brand-logo")
def brand_logo():
    path = os.path.join(BASE_DIR, LOGO_FILENAME)
    if os.path.exists(path):
        return send_from_directory(BASE_DIR, LOGO_FILENAME)
    return Response(b"GIF89a\x01\x00\x01\x00\x80\x00\x00\xff\xff\xff\x00\x00\x00!\xf9\x04\x01\n\x00\x01\x00,\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02D\x01\x00;", mimetype="image/gif")

@app.route("/__unread")
@login_required
def __unread():
    u = current_user()
    n = 0
    if u:
        db = get_db(); cur = db.cursor()
        cur.execute("SELECT COUNT(*) FROM notifications WHERE user_id=? AND is_read=0", (u["id"],))
        n = cur.fetchone()[0] or 0
        db.close()
    return "<pre>logged_in={} role={} unread={}</pre>".format(bool(u), u["role"] if u else "-", n)

# --------------------------------- Auth --------------------------------------
@app.route("/login", methods=["GET","POST"])
def login():
    if request.method=="POST":
        email = request.form.get("email","").strip().lower()
        passcode = request.form.get("passcode","").strip()
        db=get_db(); cur=db.cursor()
        cur.execute("SELECT * FROM users WHERE email=?", (email,))
        u = cur.fetchone(); db.close()
        if u and check_password_hash(u["passcode"], passcode):
            session.clear()
            session["user_id"] = u["id"]
            flash("Welcome!","message")
            return redirect(url_for("dashboard"))
        flash("Invalid email or passcode.","error")
    token = generate_csrf()
    body = """
    <div class="card" style="max-width:520px;margin:48px auto;text-align:center">
      <img src="{}" alt="logo" style="height:60px;margin-bottom:10px" onerror="this.style.display='none'">
      <h2 style="margin:6px 0">Sign in</h2>
      <form method="post" style="text-align:left">
        <input type="hidden" name="csrf_token" value="{}">
        <label>Email</label><input name="email" required>
        <label>Passcode</label><input name="passcode" type="password" required>
        <div style="margin-top:10px"><button class="btn">Login</button> <a class="btn light" href="{}">Forgot?</a></div>
      </form>
    </div>
    """.format(url_for('brand_logo'), token, url_for('forgot_password'))
    return render_page("Login", body)

@app.route("/logout")
def logout():
    session.clear(); return redirect(url_for("login"))

@app.route("/forgot", methods=["GET","POST"])
def forgot_password():
    if request.method=="POST":
        # Defensive: ensure columns exist even if an old DB slipped through
        try:
            ensure_column('password_resets', 'token', 'TEXT')
            ensure_column('password_resets', 'expires_at', 'TEXT')
        except Exception:
            pass

        email = request.form.get("email","").strip().lower()
        now = datetime.datetime.utcnow()
        token = secrets.token_urlsafe(24)
        expires = (now + datetime.timedelta(hours=6)).isoformat()
        db=get_db(); cur=db.cursor()
        cur.execute("SELECT 1 FROM users WHERE email=?", (email,))
        if cur.fetchone():
            cur.execute("INSERT INTO password_resets(user_email,state,created_at,token,expires_at) VALUES(?,?,?,?,?)",
                        (email,"open",now.isoformat(),token,expires))
            db.commit()
            reset_msg = (
                "<p>If you requested a reset, an admin will set a new passcode.</p>"
                "<p>Request token (for admin reference): <b>{}</b></p>".format(token)
            )
            send_email(email, "HMS Reset Request", reset_msg)
        db.close()
        flash("If the email exists, a reset request has been created.","message")
        return redirect(url_for("login"))
    token = generate_csrf()
    body="""
    <div class="card" style="max-width:420px;margin:48px auto">
      <h3>Forgot Passcode</h3>
      <form method="post">
        <input type="hidden" name="csrf_token" value="{}">
        <label>Your Email</label><input name="email" required>
        <div style="margin-top:10px"><button class="btn">Create Reset Request</button>
        <a class="btn light" href="{}">Back</a></div>
      </form>
    </div>
    """.format(token, url_for('login'))
    return render_page("Forgot Password", body)

@app.route("/profile", methods=["GET","POST"])
@login_required
def profile():
    u=current_user()
    if request.method=="POST":
        old=request.form.get("old",""); new=request.form.get("new","")
        if not new or len(new)<4: flash("New passcode must be at least 4 chars.","error")
        else:
            db=get_db(); cur=db.cursor()
            cur.execute("SELECT passcode FROM users WHERE id=?", (u["id"],))
            if not check_password_hash(cur.fetchone()["passcode"], old):
                db.close(); flash("Old passcode incorrect.","error"); return redirect(url_for("profile"))
            cur.execute("UPDATE users SET passcode=? WHERE id=?", (generate_password_hash(new),u["id"]))
            db.commit(); db.close(); flash("Passcode changed.","message"); return redirect(url_for("dashboard"))
    token = generate_csrf()
    body="""
    <div class="card" style="max-width:480px;margin:0 auto">
      <h3>My Profile</h3>
      <p><span class="tag">{}</span> &nbsp; <span class="tag">{}</span> &nbsp; <span class="tag">{}</span></p>
      <form method="post">
        <input type="hidden" name="csrf_token" value="{}">
        <label>Old Passcode</label><input name="old" type="password" required>
        <label>New Passcode</label><input name="new" type="password" required>
        <div style="margin-top:10px"><button class="btn">Update</button></div>
      </form>
    </div>
    """.format(u['name'], u['email'], u['role'], token)
    return render_page("Profile", body)

# -------------------------------- Dashboard ----------------------------------
@app.route("/")
@login_required
def dashboard():
    u = current_user()
    db = get_db(); cur = db.cursor()

    q_status = (request.args.get("status") or "").strip()
    q_post = (request.args.get("post") or "").strip()
    q_region = (request.args.get("region") or "").strip()
    q_from = (request.args.get("from") or "").strip()
    q_to = (request.args.get("to") or "").strip()

    restrict_hr = (u["role"]=="hr") and (u["email"].lower()!="jobs@dcdc.co.in")
    where = ["1=1"]; args=[]
    if restrict_hr: where.append("created_by=?"); args.append(u["id"])
    if u["role"]==ROLE_MANAGER: where.append("manager_owner=?"); args.append(u["id"])
    if u["role"]==ROLE_INTERVIEWER: where.append("interviewer_id=?"); args.append(u["id"])

    if q_status:
        if q_status.lower()=="joined":
            where.append("hr_join_status='joined'")
        elif q_status.lower() in ("selected","rejected"):
            where.append("lower(final_decision)=?"); args.append(q_status.lower())
        else:
            where.append("status=?"); args.append(q_status)
    if q_post: where.append("post_applied=?"); args.append(q_post)
    if q_region: where.append("assigned_region=?"); args.append(q_region)
    if q_from: where.append("date(substr(created_at,1,10))>=date(?)"); args.append(q_from)
    if q_to: where.append("date(substr(created_at,1,10))<=date(?)"); args.append(q_to)

    WHERE = " AND ".join(where)

    def scalar(sql, a=()):
        cur.execute(sql, a); r = cur.fetchone(); return r[0] if r else 0

    total = scalar(f"SELECT COUNT(*) FROM candidates WHERE {WHERE}", args)
    selected = scalar(f"SELECT COUNT(*) FROM candidates WHERE {WHERE} AND lower(final_decision)='selected'", args)
    rejected = scalar(f"SELECT COUNT(*) FROM candidates WHERE {WHERE} AND lower(final_decision)='rejected'", args)
    assigned = scalar(f"SELECT COUNT(*) FROM candidates WHERE {WHERE} AND status='Assigned'", args)
    joined = scalar(f"SELECT COUNT(*) FROM candidates WHERE {WHERE} AND hr_join_status='joined'", args)

    cur.execute("SELECT DISTINCT post_applied FROM candidates ORDER BY post_applied")
    posts = [r[0] for r in cur.fetchall() if r[0]]
    cur.execute("SELECT DISTINCT assigned_region FROM candidates WHERE assigned_region IS NOT NULL AND assigned_region<>'' ORDER BY assigned_region")
    regions = [r[0] for r in cur.fetchall()]

    cur.execute(f"""SELECT id,full_name,post_applied,status,final_decision,hr_join_status,created_at,assigned_region
                    FROM candidates WHERE {WHERE}
                    ORDER BY datetime(created_at) DESC LIMIT 20""", args)
    recent = cur.fetchall()
    recent_rows = "".join([
        (
            f"<tr><td>{h(r['full_name'])}</td><td>{h(r['post_applied'])}</td>"
            f"<td><span class='tag'>{h(r['status'])}</span></td>"
            f"<td>{h(r['final_decision'] or '-')}</td><td>{h(r['hr_join_status'] or '-')}</td>"
            f"<td>{h(r['assigned_region'] or '-')}</td>"
            f"<td>{h((r['created_at'] or '')[:19].replace('T',' '))}</td></tr>"
        )
        for r in recent
    ]) or "<tr><td colspan=7>No candidates match your filters.</td></tr>"



    cur.execute(f"SELECT COALESCE(final_decision,'(no final)') k, COUNT(*) c FROM candidates WHERE {WHERE} GROUP BY k ORDER BY c DESC", args)
    status_rows = cur.fetchall()
    cur.execute(f"SELECT COALESCE(NULLIF(assigned_region,''),'(Unassigned)') k, COUNT(*) c FROM candidates WHERE {WHERE} GROUP BY k ORDER BY c DESC", args)
    region_rows = cur.fetchall()

    cur.execute(f"""SELECT strftime('%Y-%m', COALESCE(finalized_at, created_at)) m,
                    SUM(CASE WHEN lower(final_decision)='selected' THEN 1 ELSE 0 END) sel
                    FROM candidates WHERE {WHERE} GROUP BY m ORDER BY m""", args)
    sel_map = { r["m"]: r["sel"] for r in cur.fetchall() if r["m"] }
    cur.execute(f"""SELECT strftime('%Y-%m', hr_joined_at) m, COUNT(*) j
                    FROM candidates WHERE {WHERE} AND hr_join_status='joined' AND hr_joined_at IS NOT NULL
                    GROUP BY m ORDER BY m""", args)
    join_map = { r["m"]: r["j"] for r in cur.fetchall() if r["m"] }

    months = sorted(set(list(sel_map.keys()) + list(join_map.keys())))[-12:]
    line_labels = months
    line_sel = [ sel_map.get(m,0) for m in months ]
    line_join = [ join_map.get(m,0) for m in months ]

    db.close()

    opts_status = "".join([f"<option value='{s}' {'selected' if q_status==s else ''}>{s or 'All'}</option>"
                           for s in ["","Pending","Assigned","reinterview","finalized","Selected","Rejected","Joined"]])
    opts_post = "<option value=''>All</option>" + "".join([
        f"<option value='{h(p)}' {'selected' if q_post==p else ''}>{h(p)}</option>" for p in posts
    ])
    opts_region = "<option value=''>All</option>" + "".join([
        f"<option value='{h(r)}' {'selected' if q_region==r else ''}>{h(r)}</option>" for r in regions
    ])
    page_css = """
    <style>
    .dash-grid{display:grid;grid-template-columns:1fr;gap:14px}
    @media(min-width:900px){ .tiles{display:grid;grid-template-columns:repeat(5,1fr);gap:12px} }
    .filter-bar{position:sticky;top:8px;z-index:5;background:#fff;border:1px solid var(--muted);border-radius:14px;padding:12px}
    .filters{display:grid;grid-template-columns:repeat(5, minmax(160px,1fr));gap:10px}
    .tile{background:#fff;border:1px solid var(--muted);border-radius:18px;padding:14px}
    .tile h4{margin:0 0 6px 0;color:#64748b}
    .tile .num{font-size:28px;font-weight:800}
    .t1{background:#eff6ff;border-color:#dbeafe}
    .t2{background:#f0fdf4;border-color:#dcfce7}
    .t3{background:#fef2f2;border-color:#fee2e2}
    .t4{background:#f5f3ff;border-color:#ede9fe}
    .t5{background:#fffbeb;border-color:#fef3c7}
    </style>
    """

    # ---- Charts
    chart_data = {
        "selected": selected,
        "rejected": rejected,
        "assigned": assigned,
        "region": {"labels":[r["k"] for r in region_rows], "counts":[r["c"] for r in region_rows]},
        "line": {"labels": line_labels, "sel": line_sel, "join": line_join}
    }
    charts_html = """
    <div class="card">
      <h3>Charts</h3>
      <div class="row">
        <div class="col">
          <h4>Selected / Rejected / Assigned (%)</h4>
          <canvas id="statusPie" height="240"></canvas>
        </div>
        <div class="col">
          <h4>Region-wise Candidates</h4>
          <canvas id="regionBar" height="240"></canvas>
        </div>
      </div>
      <div class="row" style="margin-top:12px">
        <div class="col">
          <h4>Selected vs Joined (Monthly)</h4>
          <canvas id="selJoinLine" height="240"></canvas>
        </div>
      </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script>
      const CD = """ + json.dumps(chart_data) + """;
      (function(){
        const statusLabels = ["Selected","Rejected","Assigned"];
        const statusCounts = [CD.selected, CD.rejected, CD.assigned];

        new Chart(document.getElementById('statusPie'), {
          type: 'pie',
          data: { labels: statusLabels, datasets: [{ data: statusCounts }] },
          options: { responsive: true }
        });

        new Chart(document.getElementById('regionBar'), {
          type: 'bar',
          data: { labels: CD.region.labels, datasets: [{ label: 'Candidates', data: CD.region.counts }] },
          options: { responsive: true, scales: { y: { beginAtZero: true } } }
        });

        new Chart(document.getElementById('selJoinLine'), {
          type: 'line',
          data: { labels: CD.line.labels, datasets: [
            { label: 'Selected', data: CD.line.sel, tension: 0.35 },
            { label: 'Joined', data: CD.line.join, tension: 0.35 }
          ]},
          options: { responsive: true }
        });
      })();
    </script>
    """

    recent_html = f"""
    <div class="card">
      <h3>Newly Added (Latest 20)</h3>
      <div class="chip">Status: {h(q_status) or 'All'}</div>
      <div class="chip">Post: {h(q_post) or 'All'}</div>
      <div class="chip">Region: {h(q_region) or 'All'}</div>
      <div class="chip">From: {h(q_from) or '—'}</div>
      <div class="chip">To: {h(q_to) or '—'}</div>
      <table>
        <thead><tr><th>Name</th><th>Post</th><th>Status</th><th>Final</th><th>Joined</th><th>Region</th><th>Created</th></tr></thead>
        <tbody>{recent_rows}</tbody>
      </table>
    </div>
    """

    filters_html = f"""
    <div class="filter-bar">
      <form method="get">
        <div class="filters">
          <div><label>Status</label><select name="status">{opts_status}</select></div>
          <div><label>Post</label><select name="post">{opts_post}</select></div>
          <div><label>Region</label><select name="region">{opts_region}</select></div>
          <div><label>From</label><input type="date" name="from" value="{h(q_from)}"></div>
          <div><label>To</label><input type="date" name="to" value="{h(q_to)}"></div>
        </div>
        <div style="margin-top:10px;display:flex;gap:8px;flex-wrap:wrap">
          <button class="btn">Apply Filters</button>
          <a class="btn light" href="{url_for('dashboard')}">Clear</a>
        </div>
      </form>
    </div>
    """

    tiles_html = f"""
    <div class="tiles">
      <div class="tile t1"><h4>Total Candidates</h4><div class="num">{total}</div></div>
      <div class="tile t2"><h4>Selected</h4><div class="num">{selected}</div></div>
      <div class="tile t3"><h4>Rejected</h4><div class="num">{rejected}</div></div>
      <div class="tile t4"><h4>Assigned</h4><div class="num">{assigned}</div></div>
      <div class="tile t5"><h4>Joined</h4><div class="num">{joined}</div></div>
    </div>
    """

    body = page_css + "<div class='dash-grid'>" + filters_html + tiles_html + charts_html + recent_html + "</div>"
    return render_page("Dashboard", body)

# ----------------------------- Notifications ---------------------------------
@app.route("/notifications")
@login_required
def notifications():
    u=current_user(); db=get_db(); cur=db.cursor()
    cur.execute("""SELECT id,title,body,is_read,created_at
                   FROM notifications WHERE user_id=? ORDER BY id DESC LIMIT 200""", (u["id"],))
    rows = cur.fetchall(); db.close()
    token = generate_csrf()
    trs = "".join([
        (
            "<tr><td>{}</td><td><strong>{}</strong><br>"
            "<div style='white-space:pre-wrap'>{}</div></td>"
            "<td>{}</td><td>"
            "<form method='post' action='{}' style='display:inline'>"
            "<input type='hidden' name='csrf_token' value='{}'>"
            "<button class='btn'>Mark read</button></form>"
            "</td></tr>"
        ).format(
            h((r['created_at'] or '')[:19].replace('T',' ')),
            h(r['title']),
            h(r['body'] or ''),
            ('Unread' if not r['is_read'] else 'Read'),
            url_for('mark_notif_read', nid=r['id']),
            token
        )
        for r in rows
    ]) or "<tr><td colspan=4>No notifications</td></tr>"
        
    body = "<div class='card'><h3>Notifications</h3><table><thead><tr><th>Time</th><th>Message</th><th>Status</th><th></th></tr></thead><tbody>{}</tbody></table></div>".format(trs)
    return render_page("Notifications", body)

@app.route("/notifications/read/<int:nid>", methods=["POST"])
@login_required
def mark_notif_read(nid):
    u=current_user(); db=get_db(); cur=db.cursor()
    cur.execute("UPDATE notifications SET is_read=1 WHERE id=? AND user_id=?", (nid, u["id"]))
    db.commit(); db.close()
    return redirect(url_for('notifications'))

# -------------------------------- Candidates ---------------------------------
@app.route("/candidates")
@login_required
def candidates_all():
    u=current_user(); db=get_db(); cur=db.cursor()

    # For Manager views: gather posts to render chips (sub-pages)
    cur.execute("SELECT DISTINCT post_applied FROM candidates ORDER BY post_applied")
    all_posts = [r[0] for r in cur.fetchall() if r[0]]

    # Filtering by post via ?post=...
    post_filter = request.args.get("post", "").strip()

    if u["role"] in (ROLE_ADMIN,ROLE_VP) or (u["role"]==ROLE_HR and is_hr_head(u)):
        if post_filter:
            cur.execute("SELECT * FROM candidates WHERE post_applied=? ORDER BY created_at DESC", (post_filter,))
        else:
            cur.execute("SELECT * FROM candidates ORDER BY created_at DESC")
    elif u["role"]==ROLE_HR:
        if post_filter:
            cur.execute("SELECT * FROM candidates WHERE created_by=? AND post_applied=? ORDER BY created_at DESC",(u["id"],post_filter))
        else:
            cur.execute("SELECT * FROM candidates WHERE created_by=? ORDER BY created_at DESC",(u["id"],))
    elif u["role"]==ROLE_MANAGER:
        if post_filter:
            cur.execute("SELECT * FROM candidates WHERE manager_owner=? AND post_applied=? ORDER BY created_at DESC",(u["id"],post_filter))
        else:
            cur.execute("SELECT * FROM candidates WHERE manager_owner=? ORDER BY created_at DESC",(u["id"],))
    else:
        if post_filter:
            cur.execute("SELECT * FROM candidates WHERE interviewer_id=? AND post_applied=? ORDER BY created_at DESC",(u["id"],post_filter))
        else:
            cur.execute("SELECT * FROM candidates WHERE interviewer_id=? ORDER BY created_at DESC",(u["id"],))

    rows = cur.fetchall(); db.close()

    def actions(r):
        role = current_user()['role']
        if role in (ROLE_MANAGER, ROLE_ADMIN):
            return ("<a class='btn light' href='{}'>Assign</a> <a class='btn' href='{}'>Finalize</a>"
                    .format(url_for('assign_candidate', candidate_id=r['id']), url_for('finalize_candidate', candidate_id=r['id'])))
        if role==ROLE_INTERVIEWER and r['interviewer_id']==current_user()['id']:
            return "<a class='btn' href='{}'>Feedback</a>".format(url_for('interview_feedback', candidate_id=r['id']))
        return "-"

    header_action = ""
    if current_user()['role'] in (ROLE_MANAGER, ROLE_ADMIN):
        header_action = "<div style='margin-bottom:10px'><a class='btn' href='{}'>Bulk Assign</a></div>".format(url_for('bulk_assign'))

    rows_html_list = []
    for r in rows:
        if r['cv_path']:
            _cv_url = url_for("download_cv", path=r["cv_path"])
            cv_html = f'<a href="{_cv_url}">CV</a>'
        else:
            cv_html = '-'

        rows_html_list.append(
            f"<tr>"
            f"<td>{h(r['candidate_code'] or '-')}</td>"
            f"<td>{h(r['full_name'])}</td>"
            f"<td>{h(r['post_applied'])}</td>"
            f"<td><span class='tag'>{h(r['status'])}</span></td>"
            f"<td>{h(r['final_decision'] or '-')}</td>"
            f"<td>{h(r['hr_join_status'] or '-')}</td>"
            f"<td>{h((r['created_at'] or '')[:19].replace('T',' '))}</td>"
            f"<td>{cv_html}</td>"
            f"<td>{actions(r)}</td>"
            "</tr>"
        )

    rows_html = "".join(rows_html_list) or "<tr><td colspan=9>No data</td></tr>"

    chips = ""
    if current_user()['role'] in (ROLE_MANAGER, ROLE_ADMIN):
        chips = "<div style='margin:6px 0'>"
        chips += '<a class="chip" href="{}">All</a>'.format(url_for('candidates_all'))
        for p in all_posts:
            chips += '<a class="chip" href="{}">{}</a>'.format(
                url_for('candidates_all') + "?" + urlencode({"post": p}),
                h(p)
            )
        if post_filter:
            chips += "<div class='badge'>Filtering by post: <b>{}</b></div>".format(h(post_filter))

    body = """
    <div class="card"><h3>All Candidates</h3>
    {}
    {}
    <table>
      <thead>
        <tr>
          <th>Candidate ID</th><th>Name</th><th>Post</th><th>Status</th>
          <th>Final</th><th>HR Join</th><th>Created</th><th>CV</th><th>Actions</th>
        </tr>
      </thead>
      <tbody>{}</tbody>
    </table>
    </div>
    """.format(header_action, chips, rows_html)
    return render_page("Candidates", body)

@app.route("/cv/<path:path>")
@login_required
def download_cv(path):
    full = os.path.abspath(os.path.join(UPLOAD_DIR, os.path.basename(path)))
    if not full.startswith(os.path.abspath(UPLOAD_DIR)) or not os.path.exists(full):
        flash("File not found.","error"); return redirect(url_for("candidates_all"))
    return send_from_directory(UPLOAD_DIR, os.path.basename(full), as_attachment=True)

# ------------------------------ Add Candidate --------------------------------
def _safe_cv_filename(name):
    base = "{}_{}".format(datetime.datetime.utcnow().strftime('%Y%m%d%H%M%S'), secrets.token_hex(4))
    ext = os.path.splitext(name.lower())[1]
    if ext not in ALLOWED_CV_EXTS: ext = ".bin"
    return base + ext

def manager_for_post(post:str):
    if post in ("Staff Nurse","Doctor","DMO"):
        return user_id_by_email("clinical_manager@dcdc.co.in")
    return user_id_by_email("dialysis.coord@dcdc.co.in")

@app.route("/add", methods=["GET","POST"])
@login_required
@role_required(ROLE_HR,ROLE_ADMIN)
def add_candidate():
    if request.method=="POST":
        f = request.form
        file = request.files.get("cv")
        cv_path=None
        if file and file.filename:
            safe = _safe_cv_filename(file.filename)
            file.save(os.path.join(UPLOAD_DIR, safe))
            cv_path=safe

        raw_phone = (f.get("phone") or "").strip()
        digits_only = "".join(ch for ch in raw_phone if ch.isdigit())
        if len(digits_only) != 10:
            flash("Mobile number must be exactly 10 digits.","error")
            return redirect(url_for("add_candidate"))

        candidate_code = (f.get("candidate_code") or "").strip() or next_candidate_code()

        fields = dict(
            candidate_code=candidate_code,
            salutation=f.get("salutation","").strip(),
            full_name=f.get("full_name","").strip(),
            email=f.get("email","").strip(),
            qualification=f.get("qualification","").strip(),
            experience_years=(f.get("experience_years") or "").strip(),
            current_designation=f.get("current_designation","").strip(),
            phone=digits_only,
            current_salary=f.get("current_salary","").strip(),
            expected_salary=f.get("expected_salary","").strip(),
            current_location=f.get("current_location","").strip(),
            preferred_location=f.get("preferred_location","").strip(),
            post_applied=f.get("post_applied","").strip(),
            interview_date=f.get("interview_date","").strip(),
            current_previous_company=f.get("current_previous_company","").strip(),
            assigned_region=f.get("assigned_region","").strip(),
            remarks=f.get("remarks","").strip(),
        )
        if not fields["full_name"] or fields["post_applied"] not in POSTS:
            flash("Name and valid Post Applied are required.","error")
            return redirect(url_for("add_candidate"))

        try:
            ey = float(fields["experience_years"]) if fields["experience_years"] else None
        except:
            ey = None

        manager_id = manager_for_post(fields["post_applied"])
        status = "Assigned"

        u=current_user(); now=datetime.datetime.utcnow().isoformat()
        db=get_db(); cur=db.cursor()
        cur.execute("""
        INSERT INTO candidates(candidate_code,salutation,full_name,email,qualification,experience_years,current_designation,phone,cv_path,current_salary,expected_salary,current_location,preferred_location,post_applied,interview_date,current_previous_company,assigned_region,status,decision_by,remarks,created_by,created_at,interviewer_id,manager_owner,final_decision,final_remark,finalized_by,finalized_at,hr_join_status,hr_joined_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,(
            fields["candidate_code"],fields["salutation"],fields["full_name"],fields["email"],fields["qualification"],ey,fields["current_designation"],fields["phone"],cv_path,fields["current_salary"],fields["expected_salary"],fields["current_location"],fields["preferred_location"],fields["post_applied"],fields["interview_date"],fields["current_previous_company"],fields["assigned_region"],status,None,fields["remarks"],u["id"],now,None,manager_id,None,None,None,None,None,None
        ))
        db.commit(); db.close()

        if manager_id:
            notify(manager_id, "Candidate Assigned to Your Role",
                   "{} (ID {}) assigned to your role.".format(fields['full_name'], candidate_code))

        flash("Candidate added (ID: {}).".format(fields['candidate_code']),"message")
        return redirect(url_for("dashboard"))

    token = generate_csrf()
    default_code = next_candidate_code()
    options="".join(["<option>{}</option>".format(p) for p in POSTS])
    body="""
    <div class="card">
      <div class="form-header">
        <h3 style="margin:0">Add Candidate</h3>
        <span class="badge">Auto-assigned to Clinical managers</span>
      </div>

      <form id="addForm" method="post" enctype="multipart/form-data">
        <input type="hidden" name="csrf_token" value="{}">
        <div class="card section blue">
          <h4>Identity & Contact</h4>
          <div class="grid-2">
            <div class="field"><label>Candidate Id</label>
              <input name="candidate_code" value="{}" placeholder="Auto-generated if left blank">
            </div>
            <div class="field"><label>Current Salary</label>
              <input name="current_salary" placeholder="₹ / month">
            </div>

            <div class="field"><label>Salutation</label>
              <input name="salutation" placeholder="Mr/Ms/Dr">
            </div>
            <div class="field"><label>Expected Salary</label>
              <input name="expected_salary" placeholder="₹ / month">
            </div>

            <div class="field"><label class="req">Name</label>
              <input name="full_name" required placeholder="Full name">
            </div>
            <div class="field"><label>Current Location</label>
              <input name="current_location" placeholder="City / State">
            </div>

            <div class="field"><label>Email</label>
              <input name="email" type="email" placeholder="name@example.com">
            </div>
            <div class="field"><label>Preferred location</label>
              <input name="preferred_location" placeholder="City / Region">
            </div>

            <div class="field"><label>Mobile No.</label>
              <input name="phone" placeholder="10 digits" type="tel" inputmode="numeric" maxlength="10"
                oninput="this.value=this.value.replace(/\\D/g,'').slice(0,10)" required>
              <small class="hint">Exactly 10 digits; numbers only. Example: 9876543210</small>
            </div>
            <div class="field"><label class="req">Post applied</label>
              <select name="post_applied">{}</select>
            </div>
          </div>
        </div>

        <div class="card section blue">
          <h4>Experience & Attachments</h4>
          <div class="grid-2">
            <div class="field"><label>Qualification</label><input name="qualification"></div>
            <div class="field"><label>Experience (years)</label><input name="experience_years" placeholder="e.g. 3.5"></div>

            <div class="field"><label>Current designation</label><input name="current_designation"></div>
            <div class="field"><label>Current/Previous company</label><input name="current_previous_company"></div>

            <div class="field"><label>Interview Date</label><input type="date" name="interview_date"></div>
            <div class="field"><label>Region</label><input name="assigned_region" placeholder="e.g. North / South"></div>

            <div class="field"><label>CV Upload (pdf/doc/docx)</label><input type="file" name="cv" accept=".pdf,.doc,.docx"></div>
            <div class="field"><label>Remarks</label><input name="remarks"></div>
          </div>
        </div>

        <div class="sticky-actions">
          <button class="btn">Save Candidate</button>
          <a class="btn light" href="{}">Cancel</a>
        </div>
      </form>
    </div>
    """.format(token, default_code, options, url_for('dashboard'))
    return render_page("Add Candidate", body)

# ----------------------- Manager: assign interviewer --------------------------
@app.route("/assign/<int:candidate_id>", methods=["GET","POST"])
@login_required
@role_required(ROLE_MANAGER, ROLE_ADMIN)
def assign_candidate(candidate_id):
    u=current_user()
    db=get_db(); cur=db.cursor()
    cur.execute("SELECT * FROM candidates WHERE id=?", (candidate_id,)); c=cur.fetchone()
    if not c: db.close(); flash("Not found.","error"); return redirect(url_for("candidates_all"))
    if u["role"]!=ROLE_ADMIN and c["manager_owner"]!=u["id"]:
        db.close(); flash("You do not own this candidate.","error"); return redirect(url_for("candidates_all"))

    if request.method=="POST":
        iid = request.form.get("interviewer_id","").strip()
        if not iid.isdigit():
            db.close(); flash("Choose an interviewer.","error"); return redirect(url_for("assign_candidate",candidate_id=candidate_id))
        cur.execute("UPDATE candidates SET interviewer_id=?, status='Assigned' WHERE id=?", (int(iid), candidate_id))
        db.commit(); db.close()

        notify(int(iid), "New Candidate Assigned",
               "{} / ID {} ({}) has been assigned to you.".format(c['full_name'], c['candidate_code'] or '-', c['post_applied']))
        if c["created_by"]:
            notify(c["created_by"], "Candidate Assigned", "{} assigned to interviewer (ID {}).".format(c['full_name'], iid))
        flash("Assigned to interviewer.","message"); return redirect(url_for("candidates_all"))

    ivs = interviewers_for_manager(u["id"]) if u["role"] != ROLE_ADMIN else all_interviewers()
    opts = "".join(["<option value='{}' {}>{}</option>".format(i['id'], "selected" if c['interviewer_id']==i['id'] else "", i['name']) for i in ivs]) or "<option disabled>No interviewers</option>"
    token = generate_csrf()
    body="""
    <div class="card" style="max-width:600px;margin:0 auto">
      <h3>Assign Interviewer</h3>
      <p><strong>{}</strong> — <span class='tag'>{}</span></p>
      <form method="post">
        <input type="hidden" name="csrf_token" value="{}">
        <label>Interviewer</label>
        <select name="interviewer_id">{}</select>
        <div style="margin-top:10px"><button class="btn">Save</button> <a class="btn light" href="{}">Back</a></div>
      </form>
    </div>
    """.format(c['full_name'], c['post_applied'], token, opts, url_for('candidates_all'))
    return render_page("Assign Interviewer", body)

@app.route("/assign/bulk", methods=["GET", "POST"])
@login_required
@role_required(ROLE_MANAGER, ROLE_ADMIN)
def bulk_assign():
    u = current_user(); db = get_db(); cur = db.cursor()

    base_where = "1=1"; args = []
    if u["role"] != ROLE_ADMIN:
        base_where = "manager_owner=?"; args = [u["id"]]

    if request.method == "GET":
        cur.execute(f"""
        SELECT id, candidate_code, full_name, post_applied, status,
        COALESCE((SELECT name FROM users uu WHERE uu.id=c.interviewer_id), '-') as current_iv
        FROM candidates c
        WHERE {base_where}
        AND (c.interviewer_id IS NULL OR c.status IN ('Assigned','reinterview'))
        ORDER BY datetime(c.created_at) DESC
        """, args)
        rows = cur.fetchall()

        ivs = interviewers_for_manager(u["id"]) if u["role"] != ROLE_ADMIN else all_interviewers()
        iv_opts = "".join([f"<option value='{i['id']}'>{i['name']}</option>" for i in ivs]) if ivs else ""

        trs = "".join([
            f"""
            <tr>
              <td><input type='checkbox' name='ids' value='{r['id']}'></td>
              <td>{h(r['candidate_code'] or '-')}</td>
              <td>{h(r['full_name'])}</td>
              <td>{h(r['post_applied'])}</td>
              <td><span class='tag'>{h(r['status'])}</span></td>
              <td>{h(r['current_iv'])}</td>
            </tr>
            """
            for r in rows
        ]) or "<tr><td colspan='6'>No candidates available for bulk assignment.</td></tr>"

        # Plain JS string (no f-string) to avoid {{ }} / arrow-function issues
        bulk_js = """
        <script>
        function ckAll(cb){
          document.querySelectorAll("input[name='ids']").forEach(el => { el.checked = cb.checked; });
        }
        </script>
        """
        token = generate_csrf()
        body = f"""
        <div class="card" style="max-width:960px;margin:0 auto">
          <h3>Bulk Assign Candidates</h3>
          <form method="post">
            <input type="hidden" name="csrf_token" value="{token}">
            <div class="row" style="margin-bottom:10px">
              <div class="col">
                <label>Assign to Interviewer</label>
                <select name="interviewer_id" required>
                  <option value="">— select —</option>
                  {iv_opts}
                </select>
              </div>
            </div>

            <table>
              <thead>
                <tr>
                  <th><input type="checkbox" onclick="ckAll(this)"></th>
                  <th>ID</th><th>Name</th><th>Post</th><th>Status</th><th>Current Interviewer</th>
                </tr>
              </thead>
              <tbody>{trs}</tbody>
            </table>

            <div class="sticky-actions">
              <button class="btn">Assign Selected</button>
              <a class="btn light" href="{url_for('candidates_all')}">Cancel</a>
            </div>
          </form>
        </div>
        """ + bulk_js
        db.close(); return render_page("Bulk Assign", body)

    iid = request.form.get("interviewer_id", "").strip()
    ids = request.form.getlist("ids")
    if not iid.isdigit() or not ids:
        db.close(); flash("Pick an interviewer and at least one candidate.", "error"); return redirect(url_for("bulk_assign"))

    placeholders = ",".join("?" for _ in ids)
    params = [int(iid)] + ids
    owner_guard = ""
    if u["role"] != ROLE_ADMIN:
        owner_guard = " AND manager_owner=?"; params.append(u["id"])

    cur.execute(f"UPDATE candidates SET interviewer_id=?, status='Assigned' WHERE id IN ({placeholders}){owner_guard}", params)

    # Notify with a detailed list
    cur.execute(f"SELECT candidate_code, full_name, post_applied FROM candidates WHERE id IN ({placeholders})", ids)
    det_rows = cur.fetchall()

    db.commit(); db.close()

    details = "\n".join([f"- {r['full_name']} (ID {r['candidate_code'] or '-'}) — {r['post_applied']}" for r in det_rows]) or "-"
    notify(int(iid), "Candidates Assigned", f"{len(det_rows)} candidates have been assigned to you:\n{details}")

    flash("Candidates assigned.","message")
    return redirect(url_for("candidates_all"))

# -------- Interviewer: feedback / edit with history (visible to manager) -----
@app.route("/interview/<int:candidate_id>", methods=["GET","POST"])
@login_required
@role_required(ROLE_INTERVIEWER)
def interview_feedback(candidate_id):
    u=current_user(); db=get_db(); cur=db.cursor()
    cur.execute("SELECT * FROM candidates WHERE id=?", (candidate_id,)); c=cur.fetchone()
    if not c or c["interviewer_id"]!=u["id"]:
        db.close(); flash("Not allowed.","error"); return redirect(url_for("candidates_all"))

    # Load last feedback by this interviewer for this candidate
    cur.execute("""SELECT * FROM interviews
                   WHERE candidate_id=? AND interviewer_id=?
                   ORDER BY id DESC LIMIT 1""", (candidate_id, u["id"]))
    last = cur.fetchone()

    if request.method=="POST":
        decision = (request.form.get("decision","") or "").strip().lower()
        rating = (request.form.get("rating","") or "").strip()
        feedback = (request.form.get("feedback","") or "").strip()
        mode = request.form.get("mode","new")  # "new" or "edit"
        try: r = int(rating)
        except: r = None
        now = datetime.datetime.utcnow().isoformat()
        if decision not in ("selected","rejected","reinterview"):
            db.close(); flash("Choose a decision.","error"); return redirect(url_for('interview_feedback',candidate_id=candidate_id))

        # If editing, keep previous row and append a new one linked via edited_from
        is_edit = 1 if (mode=="edit" and last) else 0
        edited_from = last["id"] if (mode=="edit" and last) else None

        cur.execute("""INSERT INTO interviews(candidate_id,interviewer_id,feedback,rating,decision,is_reinterview,is_edit,edited_from,created_at)
                       VALUES(?,?,?,?,?,?,?,?,?)""",
                    (candidate_id,u["id"],feedback,r,decision,1 if decision=="reinterview" else 0,is_edit,edited_from,now))

        # Update candidate status based on decision (keep 'Assigned' unless reinterview)
        if decision=="reinterview":
            cur.execute("UPDATE candidates SET status='reinterview' WHERE id=?", (candidate_id,))
        else:
            cur.execute("UPDATE candidates SET status='Assigned' WHERE id=?", (candidate_id,))
        db.commit(); db.close()

        # Notify manager with context
        if c["manager_owner"]:
            if is_edit and last:
                body = "{}: EDITED feedback.\nOld: [{} | rating {}]\nNew: [{} | rating {}]".format(
                    c['full_name'], (last['decision'] or '').upper(), last['rating'] or '-',
                    decision.upper(), r or '-'
                )
            else:
                body = "{}: {} (rating: {})".format(c['full_name'], decision.upper(), r or '-')
            notify(c["manager_owner"], "Interview Feedback Submitted", body)
        flash("Feedback {}.".format("updated" if is_edit else "submitted"),"message"); return redirect(url_for("candidates_all"))

    # Show form + previous feedback list
    history_html = ""
    cur.execute("""SELECT i.*, u.name AS iv_name
                   FROM interviews i JOIN users u ON u.id=i.interviewer_id
                   WHERE i.candidate_id=? ORDER BY i.id DESC LIMIT 10""",(candidate_id,))
    hist = cur.fetchall()
    if hist:
        items = []
        for row in hist:
            tag = "EDIT" if row["is_edit"] else ("RE-INT" if row["is_reinterview"] else "NEW")
            items.append(
                "<div class='card'><b>{}</b> — {} &nbsp; <span class='tag'>{}</span><br>"
                "Rating: {}<br><div style='white-space:pre-wrap'>{}</div></div>".format(
                    h(row["iv_name"]), h(row["decision"]), h(tag),
                    h(row["rating"] or '-'),
                    h((row["feedback"] or '').strip() or '-')
                )
            )

        history_html = "<div>{}</div>".format("".join(items))

    edit_toggle = ""
    if last:
        edit_toggle = "<div class='badge'>You have previous feedback; you may edit it. Use the selector below.</div>"

    token = generate_csrf()
    body = """
    <div class="card" style="max-width:760px;margin:0 auto">
      <h3>Interviewer Feedback</h3>
      <p><strong>{}</strong> — <span class='tag'>{}</span></p>
      {}
      <form method="post">
        <input type="hidden" name="csrf_token" value="{}">
        <div class="row">
          <div class="col"><label>Rating (1-5)</label><input name="rating" placeholder="e.g. 4"></div>
        </div>
        <label>Decision</label>
        <select name="decision">
          <option value="selected">Selected</option>
          <option value="rejected">Rejected</option>
          <option value="reinterview">Ask Re-Interview</option>
        </select>
        <label>Remarks</label>
        <textarea name="feedback" rows="5" placeholder="Notes for manager"></textarea>

        <label>Mode</label>
        <select name="mode">
          <option value="new">New feedback</option>
          <option value="edit">Edit my last feedback</option>
        </select>

        <div style="margin-top:10px"><button class="btn">Submit</button> <a class="btn light" href="{}">Back</a></div>
      </form>
    </div>
    <div class="card">
      <h4>Previous Decisions (latest first)</h4>
      {}
    </div>
    """.format(c['full_name'], c['post_applied'], edit_toggle, token, url_for('candidates_all'), history_html)
    db.close()
    return render_page("Interviewer Feedback", body)

# --------------------------- Finalize / HR join -------------------------------
@app.route("/finalize/<int:candidate_id>", methods=["GET","POST"])
@login_required
@role_required(ROLE_MANAGER,ROLE_ADMIN)
def finalize_candidate(candidate_id):
    u=current_user(); db=get_db(); cur=db.cursor()
    cur.execute("SELECT * FROM candidates WHERE id=?", (candidate_id,)); c=cur.fetchone()
    if not c: db.close(); flash("Not found.","error"); return redirect(url_for("dashboard"))
    if u["role"]!=ROLE_ADMIN and c["manager_owner"]!=u["id"]:
        db.close(); flash("You do not own this candidate.","error"); return redirect(url_for("dashboard"))

    # Show last two interviews for context (to see edits)
    cur.execute("""SELECT i.*,u.name interviewer_name
                   FROM interviews i JOIN users u ON u.id=i.interviewer_id
                   WHERE i.candidate_id=? ORDER BY i.id DESC LIMIT 2""",(candidate_id,))
    last2 = cur.fetchall()
    last_block = "<p>No interview yet.</p>"
    if last2:
        cards=[]
        for idx,row in enumerate(last2, start=1):
            tag = "EDIT" if row["is_edit"] else ("RE-INT" if row["is_reinterview"] else "NEW")
            cards.append(
                "<div class='card'><strong>{} #{}</strong><br>By: {}<br>Rating: {} / 5<br>"
                "Decision: {} <span class='tag'>{}</span><br>"
                "Notes:<div style='white-space:pre-wrap'>{}</div></div>".format(
                    "Entry", idx,
                    h(row['interviewer_name']),
                    h(row['rating'] or '-'),
                    h(row['decision']),
                    h(tag),
                    h((row['feedback'] or '').strip() or '-')
                )
            )

        last_block = "".join(cards)

    if request.method=="POST":
        action=request.form.get("action"); remark=request.form.get("remark","").strip()
        now=datetime.datetime.utcnow().isoformat()
        if action == "select":
            cur.execute("""
            UPDATE candidates
            SET status='finalized',
                final_decision='selected',
                final_remark=?,
                finalized_by=?,
                finalized_at=?,
                decision_by=?,
                interviewer_id=NULL
            WHERE id=?""",
            (remark, u["id"], now, u["id"], candidate_id))
        elif action == "reject":
            cur.execute("""
            UPDATE candidates
            SET status='finalized',
                final_decision='rejected',
                final_remark=?,
                finalized_by=?,
                finalized_at=?,
                decision_by=?,
                interviewer_id=NULL
            WHERE id=?""",
            (remark, u["id"], now, u["id"], candidate_id))
        elif action=="reinterview":
            cur.execute("UPDATE candidates SET status='reinterview', final_decision=NULL, final_remark=?, interviewer_id=NULL WHERE id=?",
                        (remark,candidate_id))
        else:
            db.close(); flash("Invalid action.","error"); return redirect(url_for("finalize_candidate",candidate_id=candidate_id))
        db.commit(); db.close()

        for uid in filter(None, [c["created_by"], c["interviewer_id"], c["manager_owner"]]):
            notify(uid, "Candidate Finalized", "{} -> {}. Remark: {}".format(c['full_name'], action.upper(), (remark or '-')))

        if action in ("select","reject"):
            title = "Candidate Selected" if action=="select" else "Candidate Rejected"
            msg = "{} (ID {}) was {} by manager.".format(c['full_name'], c['candidate_code'] or '-', "SELECTED" if action=="select" else "REJECTED")
            db2 = get_db(); cur2 = db2.cursor()
            cur2.execute("SELECT id FROM users WHERE role='hr'")
            hr_ids = [row["id"] for row in cur2.fetchall()]
            db2.close()
            for hid in hr_ids: notify(hid, title, msg)

        flash("Final decision updated.","message"); return redirect(url_for("dashboard"))

    token = generate_csrf()
    body="""
    <div class="card" style="max-width:720px;margin:0 auto">
      <h3>Finalize Candidate</h3>
      <p><strong>{}</strong> — <span class='tag'>{}</span></p>
      {}
      <form method="post">
        <input type="hidden" name="csrf_token" value="{}">
        <label>Final Remark</label><textarea name="remark" rows="4"></textarea>
        <div style="margin-top:10px;display:flex;gap:8px;flex-wrap:wrap">
          <button name="action" value="select" class="btn">Select</button>
          <button name="action" value="reject" class="btn danger">Reject</button>
          <button name="action" value="reinterview" class="btn warn">Re-Interview</button>
          <a class="btn light" href="{}">Cancel</a>
        </div>
      </form>
    </div>
    """.format(c['full_name'], c['post_applied'], last_block, token, url_for('dashboard'))
    return render_page("Finalize", body)

@app.route("/hr/queue")
@login_required
@role_required(ROLE_HR,ROLE_ADMIN)
def hr_join_queue():
    u = current_user(); db = get_db(); cur = db.cursor()

    base_sql = """
    SELECT c.id, c.full_name, c.post_applied,
           COALESCE(c.final_remark,'-') AS final_remark,
           strftime('%Y-%m-%d %H:%M', c.finalized_at) AS finalized_at,
           mu.name AS finalized_by_name
    FROM candidates c
    LEFT JOIN users mu ON mu.id = c.finalized_by
    WHERE c.status='finalized'
      AND lower(c.final_decision)='selected'
      AND c.hr_join_status IS NULL
    """

    if is_hr_head(u) or u["role"]==ROLE_ADMIN:
        cur.execute(base_sql + " ORDER BY c.finalized_at DESC")
    else:
        cur.execute(base_sql + " AND c.created_by=? ORDER BY c.finalized_at DESC", (u["id"],))

    rows = cur.fetchall(); db.close()

    if not rows:
        body = "<div class='card'><h3>Awaiting Join Status</h3><ul><li>None</li></ul></div>"
        return render_page("HR Actions", body)

    trs = "".join([
        f"""
        <tr>
          <td>{h(r['full_name'])}</td>
          <td>{h(r['post_applied'])}</td>
          <td>{h(r['finalized_by_name'] or '-')}</td>
          <td style="white-space:pre-wrap">{h(r['final_remark'])}</td>
          <td>{h(r['finalized_at'] or '-')}</td>
          <td><a class="btn" href="{url_for('hr_join_update', candidate_id=r['id'])}">Mark Join</a></td>
        </tr>
        """
        for r in rows
    ]) or "<tr><td colspan=6>No candidates found.</td></tr>"

    body = """
    <div class="card">
      <h3>Awaiting Join Status</h3>
      <table>
        <thead>
          <tr><th>Name</th><th>Post</th><th>Finalized By</th><th>Final Remark</th><th>Finalized At</th><th>Action</th></tr>
        </thead>
        <tbody>{}</tbody>
      </table>
    </div>
    """.format(trs)
    return render_page("HR Actions", body)

@app.route("/hr/join/<int:candidate_id>", methods=["GET","POST"])
@login_required
@role_required(ROLE_HR,ROLE_ADMIN)
def hr_join_update(candidate_id):
    u=current_user(); db=get_db(); cur=db.cursor()
    cur.execute("SELECT * FROM candidates WHERE id=?", (candidate_id,)); c=cur.fetchone()
    if not c: db.close(); flash("Not found.","error"); return redirect(url_for("hr_join_queue"))
    if u["role"]==ROLE_HR and not is_hr_head(u) and c["created_by"]!=u["id"]:
        db.close(); flash("You cannot edit another HR's candidate.","error"); return redirect(url_for("hr_join_queue"))
    if c["final_decision"]!="selected" or c["status"]!="finalized":
        db.close(); flash("Only finalized 'Selected' candidates are updatable.","error"); return redirect(url_for("hr_join_queue"))

    cur.execute("SELECT name FROM users WHERE id=?", (c["manager_owner"],)); row_mgr = cur.fetchone()
    manager_name = row_mgr["name"] if row_mgr else "-"
    cur.execute("SELECT name FROM users WHERE id=?", (c["finalized_by"],)); row_fin = cur.fetchone()
    finalized_by_name = row_fin["name"] if row_fin else "-"

    if request.method == "POST":
        st = request.form.get("status")
        reason = request.form.get("reason", "").strip() if st == "not_joined" else None

        if st not in ("joined", "not_joined"):
            db.close(); flash("Invalid status.","error"); return redirect(url_for("hr_join_update", candidate_id=candidate_id))
        if st == "not_joined" and not reason:
            db.close(); flash("Please provide reason for Not Joined.","error"); return redirect(url_for("hr_join_update", candidate_id=candidate_id))

        now = datetime.datetime.utcnow().isoformat()
        cur.execute(
            "UPDATE candidates SET hr_join_status=?, hr_joined_at=?, status='closed', final_remark=? WHERE id=?",
            (st, now, reason if reason else c["final_remark"], candidate_id)
        )
        db.commit(); db.close()

        msg = "{} join status: {}".format(c['full_name'], st.upper()) + ((" (Reason: {})".format(reason)) if reason else "")
        for uid in filter(None, [c["manager_owner"], c["finalized_by"], c["created_by"]]):
            notify(uid, "Join Status Updated", msg)

        flash("Join status updated.","message")
        return redirect(url_for("dashboard"))

    token = generate_csrf()
    body = """
    <div class="card" style="max-width:700px;margin:0 auto">
      <h3>HR: Mark Join Status</h3>
      <div class="card section blue">
        <h4>Candidate</h4>
        <p><strong>{}</strong> — <span class='tag'>{}</span></p>
        <p>Manager: <strong>{}</strong></p>
        <p>Finalized By: <strong>{}</strong></p>
        <p>Final Remark: <span style="white-space:pre-wrap">{}</span></p>
      </div>

      <form method="post">
        <input type="hidden" name="csrf_token" value="{}">
        <label>Status</label>
        <select name="status" id="status" onchange="toggleReason()">
          <option value="joined">Joined</option>
          <option value="not_joined">Not Joined</option>
        </select>

        <div id="reasonBox" style="margin-top:10px; display:none;">
          <label>Reason (if Not Joined)</label>
          <textarea name="reason" rows="3" placeholder="Explain why the candidate did not join"></textarea>
        </div>

        <div style="margin-top:10px">
          <button class="btn">Save</button>
          <a class="btn light" href="{}">Back</a>
        </div>
      </form>
    </div>
    <script>
    function toggleReason(){ var st = document.getElementById('status').value;
      document.getElementById('reasonBox').style.display = (st === 'not_joined') ? 'block' : 'none'; }
    </script>
    """.format(c['full_name'], c['post_applied'], manager_name, finalized_by_name, (c['final_remark'] or '-'), token, url_for('hr_join_queue'))
    return render_page("HR Join Update", body)

# ---------------------------------- Admin ------------------------------------
@app.route("/admin/users", methods=["GET","POST"])
@login_required
@role_required(ROLE_ADMIN)
def admin_users():
    db=get_db(); cur=db.cursor()
    if request.method=="POST":
        name=request.form.get("name","").strip()
        email=request.form.get("email","").strip().lower()
        role=request.form.get("role","").strip()
        manager_id=request.form.get("manager_id","").strip()
        passcode=request.form.get("passcode","").strip()
        if not name or not email or role not in (ROLE_ADMIN,ROLE_VP,ROLE_HR,ROLE_MANAGER,ROLE_INTERVIEWER) or not passcode:
            flash("Provide name, email, role, passcode.","error")
        else:
            mid=int(manager_id) if manager_id.isdigit() else None
            try:
                cur.execute("INSERT INTO users(name,email,role,manager_id,passcode,created_at) VALUES(?,?,?,?,?,?)",
                            (name,email,role,mid,generate_password_hash(passcode),datetime.datetime.utcnow().isoformat()))
                db.commit(); flash("User added.","message")
            except sqlite3.IntegrityError:
                flash("Email already exists.","error")
    cur.execute("SELECT id,name,email,role,manager_id FROM users ORDER BY role,name")
    users=cur.fetchall()
    opts_role="".join(["<option value='{}'>{}</option>".format(r, r) for r in [ROLE_ADMIN,ROLE_VP,ROLE_HR,ROLE_MANAGER,ROLE_INTERVIEWER]])
    cur.execute("SELECT id,name FROM users WHERE role IN ('manager') ORDER BY name"); mgrs=cur.fetchall()
    opts_mgr="<option value=''>—</option>" + "".join(["<option value='{}'>{}</option>".format(m['id'], m['name']) for m in mgrs])
    rows="".join(["<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>".format(u['id'],u['name'],u['email'],u['role'],u['manager_id'] or '-') for u in users])
    db.close()
    token = generate_csrf()
    body="""
    <div class="card">
      <h3>Add User</h3>
      <form method="post">
        <input type="hidden" name="csrf_token" value="{}">
        <div class="row">
          <div class="col"><label>Name</label><input name="name" required></div>
          <div class="col"><label>Email</label><input name="email" required></div>
          <div class="col"><label>Role</label><select name="role">{}</select></div>
          <div class="col"><label>Manager (if interviewer)</label><select name="manager_id">{}</select></div>
          <div class="col"><label>Passcode</label><input name="passcode" required></div>
        </div>
        <div style="margin-top:10px"><button class="btn">Create</button></div>
      </form>
    </div>
    <div class="card">
      <h3>Users</h3>
      <table><thead><tr><th>ID</th><th>Name</th><th>Email</th><th>Role</th><th>Manager</th></tr></thead>
      <tbody>{}</tbody></table>
    </div>
    <div class="card">
      <h3>Password Reset Requests</h3>
      <p><a class="btn" href="{}">Manage Resets</a></p>
    </div>
    """.format(token, opts_role, opts_mgr, rows or '<tr><td colspan=5>No users</td></tr>', url_for('admin_resets'))
    return render_page("Admin: Users", body)

@app.route("/admin/resets", methods=["GET","POST"])
@login_required
@role_required(ROLE_ADMIN)
def admin_resets():
    db=get_db(); cur=db.cursor()
    if request.method=="POST":
        rid = request.form.get("rid","").strip()
        newp = request.form.get("new","").strip()
        tok  = request.form.get("token","").strip()
        
        if rid.isdigit() and len(newp) >= 4 and tok:
            cur.execute("""SELECT * FROM password_resets
                           WHERE id=? AND state='open' AND token=?
                             AND datetime(expires_at) > datetime('now')""",
                        (int(rid), tok))


            row=cur.fetchone()
            if row:
                cur.execute("UPDATE users SET passcode=? WHERE email=?", (generate_password_hash(newp),row["user_email"]))
                cur.execute("""UPDATE password_resets SET state='resolved', resolved_at=?, resolver_id=?, new_passcode=NULL WHERE id=?""",
                            (datetime.datetime.utcnow().isoformat(), current_user()["id"], int(rid)))
                db.commit();
                try:
                    html = (
                        "<p>Your passcode has been reset by an administrator.</p>"
                        "<p>Temporary passcode: <b>{}</b></p>"
                        "<p>Please sign in and change it immediately.</p>"
                    ).format(newp)
                    send_email(row["user_email"], "HMS Passcode Reset", html)
                except Exception:
                    pass
                flash("Reset resolved and passcode updated.","message")
            else:
                flash("Reset not found or already resolved.","error")
        else:
            flash("Provide valid request ID and a new passcode (>=4 chars).","error")

    cur.execute("SELECT * FROM password_resets ORDER BY created_at DESC")
    rows=cur.fetchall(); db.close()
    def tr(r):
        return "<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>".format(
            r['id'], r['user_email'], r['state'], (r['created_at'] or '')[:19].replace('T',' '),
            (r['resolved_at'] or '')[:19].replace('T',' ') if r['resolved_at'] else '-'
        )
    table="".join([tr(r) for r in rows]) or "<tr><td colspan=5>No requests</td></tr>"
    token = generate_csrf()
    body="""
    <div class="card">
      <h3>Password Reset Requests</h3>
      <table><thead><tr><th>ID</th><th>Email</th><th>State</th><th>Created</th><th>Resolved</th></tr></thead>
      <tbody>{}</tbody></table>
    </div>
    <div class="card" style="max-width:560px">
      <h3>Resolve a Request</h3>
      <form method="post">
          <input type="hidden" name="csrf_token" value="{}">
          <label>Request ID</label><input name="rid" required>
          <label>Token</label><input name="token" required>
          <label>New Passcode</label><input name="new" required>
          <div style="margin-top:10px">
            <button class="btn">Set New Passcode</button>
            <a class="btn light" href="{}">Back</a>
          </div>
        </form>

    </div>
    """.format(table, token, url_for('admin_users'))
    return render_page("Admin: Reset Requests", body)

# --------------------------------- Bulk upload --------------------------------
@app.route("/bulk/sample")
@login_required
@role_required(ROLE_HR,ROLE_ADMIN)
def bulk_sample():
    headers = ["Candidate Id","Salutation","Name","Email","Qualification","Experience (years)","Current designation",
               "Mobile No.","Current Salary","Expected Salary","Current Location","Preferred location","Post applied",
               "Interview Date","Current/Previous company","Region","Status","remarks"]
    wb = Workbook(); ws = wb.active; ws.title = "Candidates"
    for i,h in enumerate(headers, start=1): ws.cell(row=1, column=i).value = h
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="bulk_sample.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/bulk", methods=["GET","POST"])
@login_required
@role_required(ROLE_HR,ROLE_ADMIN)
def bulk_upload():
    if request.method=="POST":
        file = request.files.get("xlsx")
        if not file or not file.filename.lower().endswith(".xlsx"):
            flash("Please upload an .xlsx file.","error"); return redirect(url_for("bulk_upload"))

        safe = "bulk_{}_{}.xlsx".format(datetime.datetime.utcnow().strftime('%Y%m%d%H%M%S'), secrets.token_hex(3))
        xpath = os.path.join(UPLOAD_DIR, safe); file.save(xpath)

        try:
            wb = load_workbook(xpath); ws = wb.active
            headers = [ (ws.cell(row=1,column=i).value or "").strip().lower() for i in range(1, ws.max_column+1) ]
            def idx(label):
                l=label.strip().lower()
                return headers.index(l) if l in headers else None

            m = { k:idx(k) for k in [
                "candidate id","salutation","name","email","qualification","experience (years)",
                "current designation","mobile no.","current salary","expected salary","current location",
                "preferred location","post applied","interview date","current/previous company","region","status","remarks"
            ]}

            inserted=0; bad_phone=0; bad_post_or_name=0
            db=get_db(); cur=db.cursor()
            now = datetime.datetime.utcnow().isoformat()
            u=current_user()

            cur.execute("SELECT MAX(id) FROM candidates"); next_base = (cur.fetchone()[0] or 0)

            for r in range(2, ws.max_row+1):
                def v(key):
                    ci = m.get(key); return (ws.cell(row=r, column=ci+1).value if ci is not None else "") or ""
                post=str(v("post applied")).strip(); full_name=str(v("name")).strip()
                if post not in POSTS or not full_name: bad_post_or_name+=1; continue

                digits = "".join(ch for ch in str(v("mobile no.")).strip() if ch.isdigit())
                if len(digits) != 10: bad_phone+=1; continue

                try: ey=float(v("experience (years)")) if str(v("experience (years)"))!="" else None
                except: ey=None

                cand_code = str(v("candidate id")).strip()
                if not cand_code: next_base += 1; cand_code = "DCDC_C{}".format(next_base)

                manager_id = manager_for_post(post); status = "Assigned"
                cur.execute("""
                INSERT INTO candidates(candidate_code,salutation,full_name,email,qualification,experience_years,current_designation,phone,cv_path,current_salary,expected_salary,current_location,preferred_location,post_applied,interview_date,current_previous_company,assigned_region,status,decision_by,remarks,created_by,created_at,interviewer_id,manager_owner,final_decision,final_remark,finalized_by,finalized_at,hr_join_status,hr_joined_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,(
                    cand_code, str(v("salutation")).strip(), full_name, str(v("email")).strip(),
                    str(v("qualification")).strip(), ey, str(v("current designation")).strip(), digits,
                    None, str(v("current salary")).strip(), str(v("expected salary")).strip(), str(v("current location")).strip(),
                    str(v("preferred location")).strip(), post, str(v("interview date")).strip(), str(v("current/previous company")).strip(),
                    str(v("region")).strip(), status, None, str(v("remarks")).strip(),
                    u["id"], now, None, manager_id, None, None, None, None, None, None
                ))
                inserted+=1

                if manager_id:
                    notify(manager_id, "Candidate Assigned to Your Role",
                           "{} (ID {}) assigned to your role.".format(full_name, cand_code))

            db.commit(); db.close()
            flash("Bulk upload complete. Inserted {}. Skipped {} (bad name/post), {} (invalid phone).".format(inserted,bad_post_or_name,bad_phone),"message")
            return redirect(url_for('candidates_all'))
        except Exception as e:
            flash("Upload failed: {}. Supported format: .xlsx".format(e),"error"); return redirect(url_for('bulk_upload'))

    sample_cols = ", ".join([
        "Candidate Id","Salutation","Name","Email","Qualification","Experience (years)","Current designation",
        "Mobile No.","Current Salary","Expected Salary","Current Location","Preferred location","Post applied",
        "Interview Date","Current/Previous company","Region","Status","remarks"
    ])
    token = generate_csrf()
    body="""
    <div class="card" style="max-width:700px;margin:0 auto">
      <h3>Bulk Upload (Excel .xlsx)</h3>
      <p>Expected columns: <span class="tag">{}</span></p>
      <p><a class="btn light" href="{}">Download Sample Excel</a></p>
      <form method="post" enctype="multipart/form-data">
        <input type="hidden" name="csrf_token" value="{}">
        <label>Choose .xlsx file</label><input type="file" name="xlsx" accept=".xlsx" required>
        <div style="margin-top:10px"><button class="btn">Upload</button> <a class="btn light" href="{}">Back</a></div>
      </form>
    </div>
    """.format(sample_cols, url_for('bulk_sample'), token, url_for('candidates_all'))
    return render_page("Bulk Upload", body)

# --------------------------------- Run (dev) ---------------------------------
if __name__=="__main__":
    port=int(os.environ.get("PORT",5000))
    app.run(host="0.0.0.0", port=port, debug=True)
